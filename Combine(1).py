import openpyxl

# create a list of files to combine
file_list = ['File 1.xlsx', 'File 2.xlsx', 'File 3.xlsx']
# create a file name for the output file
output_file = 'combined_files.xlsx'


#extract the data from the files and put it in a list
def extract_data(file_name):
    # open the file
    wb = openpyxl.load_workbook(file_name)
    # get the active sheet
    sheet = wb.active
    # create a list to store the data
    data = []
    # loop through the rows
    for row in sheet.iter_rows():
        # create a list to store the data from each row
        row_data = []
        # loop through the columns
        for cell in row:
            # add the cell's data to the row_data list
            row_data.append(cell.value)
        # add the row_data list to the data list
        data.append(row_data)
    # return the data
    return data


x = extract_data(file_list[0])
y = extract_data(file_list[1])
z = extract_data(file_list[2])

title = x[0][0]
title2 = x[0][1]
title3 = x[0][2]
title4 = x[0][3]

def search_row(data, title):
    for row in data:
        if row[0] == title:
            return int(0)
        if row[1] == title:
            return int(1)
        if row[2] == title:
            return int(2)
        if row[3] == title:
            return int(3) 
    return None

# get the column values from the rows numbers in the list
def get_column(data, row_num):
    column = []
    for row in data:
        column.append(row[row_num])
    return column


# create a new excel file 
wb = openpyxl.Workbook()
# get the active sheet
sheet = wb.active
# get list of all row in the sheet
rows = sheet.rows
# write only to the first cell in the first row
sheet['A1'] = title
# write only to the first cell in the second row
sheet['B1'] = title2
# write only to the first cell in the third row
sheet['C1'] = title3
# write only to the first cell in the fourth row
sheet['D1'] = title4

count = 0
# write the data from the first file to the sheet
try : 
    for i in range(len(get_column(x, search_row(x, 'Student ID'))[1:])):
        sheet.cell(row=i+2, column=1).value = get_column(x, search_row(x, 'Student ID'))[i+1]
        sheet.cell(row=i+2, column=2).value = get_column(x, search_row(x, 'Name'))[i+1]
        sheet.cell(row=i+2, column=3).value = get_column(x, search_row(x, 'Location'))[i+1]
        sheet.cell(row=i+2, column=4).value = get_column(x, search_row(x, 'Courses'))[i+1]
        count += 1
except TypeError:
    for i in range(len(get_column(x, search_row(x, 'Student ID '))[1:])):
        sheet.cell(row=i+2, column=1).value = get_column(x, search_row(x, 'Student ID '))[i+1]
        sheet.cell(row=i+2, column=2).value = get_column(x, search_row(x, 'Name'))[i+1]
        sheet.cell(row=i+2, column=3).value = get_column(x, search_row(x, 'Location'))[i+1]
        sheet.cell(row=i+2, column=4).value = get_column(x, search_row(x, 'Courses'))[i+1]
        count += 1
# write the data from the second file to the sheet start from the last row
try:
    for i in range(len(get_column(y, search_row(y, 'Student ID'))[1:])):
        sheet.cell(row=i+count+2, column=1).value = get_column(y, search_row(y, 'Student ID'))[i+1]
        sheet.cell(row=i+count+2, column=2).value = get_column(y, search_row(y, 'Name'))[i+1]
        sheet.cell(row=i+count+2, column=3).value = get_column(y, search_row(y, 'Location'))[i+1]
        sheet.cell(row=i+count+2, column=4).value = get_column(y, search_row(y, 'Courses'))[i+1]
        count += 1
except TypeError:
    for i in range(len(get_column(y, search_row(y, 'Student ID '))[1:])):
        sheet.cell(row=i+count+2, column=1).value = get_column(y, search_row(y, 'Student ID '))[i+1]
        sheet.cell(row=i+count+2, column=2).value = get_column(y, search_row(y, 'Name'))[i+1]
        sheet.cell(row=i+count+2, column=3).value = get_column(y, search_row(y, 'Location'))[i+1]
        sheet.cell(row=i+count+2, column=4).value = get_column(y, search_row(y, 'Courses'))[i+1]
        count += 1

# write the data from the third file to the sheet start from the last row
try:
    for i in range(len(get_column(z, search_row(z, 'Student ID'))[1:])):
        sheet.cell(row=i+count+2, column=1).value = get_column(z, search_row(z, 'Student ID'))[i+1]
        sheet.cell(row=i+count+2, column=2).value = get_column(z, search_row(z, 'Name'))[i+1]
        sheet.cell(row=i+count+2, column=3).value = get_column(z, search_row(z, 'Location'))[i+1]
        sheet.cell(row=i+count+2, column=4).value = get_column(z, search_row(z, 'Courses'))[i+1]
        count += 1
except TypeError:
    for i in range(len(get_column(z, search_row(z, 'Student ID '))[1:])):
        sheet.cell(row=i+count+2, column=1).value = get_column(z, search_row(z, 'Student ID '))[i+1]
        sheet.cell(row=i+count+2, column=2).value = get_column(z, search_row(z, 'Name'))[i+1]
        sheet.cell(row=i+count+2, column=3).value = get_column(z, search_row(z, 'Location'))[i+1]
        sheet.cell(row=i+count+2, column=4).value = get_column(z, search_row(z, 'Courses'))[i+1]
        count += 1

#check every row in the sheet and if the value is empty then delete the row
for row in sheet.iter_rows():
    for cell in row:
        if cell.value == None:
            sheet.delete_rows(cell.row)
            break


wb.save(output_file)