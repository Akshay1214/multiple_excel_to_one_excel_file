# Imoprting Libraries 
#from pathlib import Path
import openpyxl as xl
import glob 
import os
from openpyxl import load_workbook, Workbook

# Loading all excel files in folder 
path = os.getcwd()
excel_files = glob.glob(os.path.join(path, "*.xlsx"))
print(excel_files)
# files_dir = r"C:/Users/Admin/Desktop/On Going/multiple_excel_to_one_excel_file-main"
# excel_files = list(Path(files_dir).glob('*.xlsx'))
# print(excel_files)

# Creates a new empty excel file to store all merged data
wb = Workbook()
ws = wb.active
ws.title = "Merged Data"
wb.save(filename = 'All_data.xlsx')

# Opening the destination excel file
output_file = "All_data.xlsx"    #"C:/Users/Admin/Desktop/On Going/multiple_excel_to_one_excel_file-main/All_data.xlsx"
wb2 = xl.load_workbook(output_file)
ws2 = wb2.active

current_row = 1
for file in excel_files:
    # Opening the source excel file
    wb1 = xl.load_workbook(file)
    ws1 = wb1.worksheets[0]

    # Calculate total number of rows and columns in source excel file
    mr = ws1.max_row
    mc = ws1.max_column

    # Copying the cell values from source excel file to destination excel file
    for i in range (1, mr + 1):
        for j in range (1, mc + 1):
            # Reading cell value from source excel file
            c = ws1.cell(row = i, column = j)

            # Writing the read value to destination excel file
            ws2.cell(row = current_row, column = j).value = c.value
        current_row += 1
        
    # Saving the destination excel file
    wb2.save(str(output_file))    





'''
# Opening the source excel file
filename ="C:/Users/Admin/Desktop/On Going/multiple_excel_to_one_excel_file-main/File 1.xlsx"
wb1 = xl.load_workbook(filename)
ws1 = wb1.worksheets[0]

# Opening the destination excel file
filename1 ="C:/Users/Admin/Desktop/On Going/multiple_excel_to_one_excel_file-main/All_data.xlsx"
wb2 = xl.load_workbook(filename1)
ws2 = wb2.active

# Calculate total number of rows and columns in source excel file
mr = ws1.max_row
mc = ws1.max_column

# Copying the cell values from source excel file to destination excel file
for i in range (1, mr + 1):
	for j in range (1, mc + 1):
		# Reading cell value from source excel file
		c = ws1.cell(row = i, column = j)

		# Writing the read value to destination excel file
		ws2.cell(row = i, column = j).value = c.value

# Saving the destination excel file
wb2.save(str(filename1))
'''



'''
import os
import pandas as pd
cwd = os.path.abspath('') 
files = os.listdir(cwd)  
df = pd.DataFrame()
for file in files:
    if file.endswith('.xlsx'):
        df = df.append(pd.read_excel(file), ignore_index=True) 
df.head() 
df.to_excel('total_sales.xlsx')

'''