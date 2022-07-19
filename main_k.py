import glob
import os
from openpyxl import load_workbook, Workbook


def read_files():
    list2 = []
    path = os.getcwd()
    excel_files = glob.glob(os.path.join(path, "*.xlsx"))

    for file in excel_files:
        wb = load_workbook(file)
        ws = wb.active

        for col_cells in ws.iter_cols(min_col=0, max_col=ws.max_column):
            lst1 = []
            for cell in col_cells:
                lst1.append(cell.value)
            list2.append(lst1)
    write_data(list2)


def iter_rows(ws):
    for row in ws.iter_rows():
        yield [cell.value for cell in row if cell.value is not None]


def write_data(list2):
    emp_id = []
    name = []
    desi = []
    loc = []
    for i in list2:
        if "Emp ID" in i:
            for temp in i:
                emp_id.append(temp)
        if "Name" in i:
            for temp in i:
                name.append(temp)
        if "Location" in i:
            for temp in i:
                loc.append(temp)
        if "Designation" in i:
            for temp in i:
                desi.append(temp)
    print(emp_id)
    print(desi)
    print(loc)
    print(name)
    write_dataa(emp_id, desi, loc, name)


def write_dataa(emp_id, desi, loc, name):
    wb = Workbook()
    ws = wb.active

    for i, statN in enumerate(emp_id):
        ws.cell(row=i + 1, column=1).value = statN

    for i, statN in enumerate(desi):
        ws.cell(row=i + 1, column=2).value = statN

    for i, statN in enumerate(loc):
        ws.cell(row=i + 1, column=3).value = statN

    for i, statN in enumerate(name):
        ws.cell(row=i + 1, column=4).value = statN

    wb.save("Output.xlsx")


if __name__ == '__main__':
    read_files()
