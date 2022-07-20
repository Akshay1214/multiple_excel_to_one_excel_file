from openpyxl import Workbook, load_workbook
import os


def compile_workbooks(workbooks_path, final_filename):
    wbs = []
    for file in os.listdir(workbooks_path):
        if not file.startswith("~$") and file.endswith(".xlsx"):
            wb = load_workbook(os.path.join(workbooks_path, file))
            wbs.append(wb)
    
    final_wb = Workbook()
    final_ws = final_wb.worksheets[0]

    wb1 = wbs[0]
    ws1 = wb1.worksheets[0] 
 
    for j in range(1, ws1.max_column+1):
        final_ws.cell(row=1, column=j).value = ws1.cell(row=1, column=j).value

    current_row = 2

    for wb in wbs:
        for ws in wb.worksheets:
            mr = ws.max_row 
            mc = ws.max_column 

            for i in range (2, mr + 1): 
                for j in range (1, mc + 1): 
                    current_cell = ws.cell(row = i, column = j) 
                    final_ws.cell(row = current_row, column = j).value = current_cell.value

                current_row += 1

    final_wb.save(os.path.join(workbooks_path, final_filename))
compile_workbooks(os.path.join(os.getcwd()), "final.xlsx")

# if __name__ == '__main__':
#     compile_workbooks(os.path.join(os.getcwd()), "final.xlsx")