import openpyxl


def exel_to_dict(file_path: str):
    work_book = openpyxl.load_workbook(file_path)
    work_sheet = work_book.active
    
    max_rows = work_sheet.max_row
    max_cols = work_sheet.max_column
    
    output_dict = {}
    for header_index in range(1, max_cols+1):
        header = work_sheet.cell(1, header_index).value.strip().replace(' ', '_').lower()
        data = []
        for index in range(2, max_rows+1):
            data.append(work_sheet.cell(index, header_index).value)
        
        output_dict[header] = data
        
    return output_dict


def MergeFiles(*args, output_file='Combined.xlsx'):
    dict_files = [exel_to_dict(file_path) for file_path in args]
    work_book = openpyxl.Workbook()
    work_sheet = work_book.active
    final_output_dict = {}

    for header in list(dict_files[0].keys()):
        data = []
        for output_dict in dict_files:
            data.extend(output_dict[header])

        final_output_dict[header] = data

    for col, header in enumerate(list(final_output_dict.keys())):
        header_cell = work_sheet.cell(1, col + 1)
        header_cell.value = header.replace('_', ' ').capitalize()

        for row, data in enumerate(final_output_dict[header]):
            data_cell = work_sheet.cell(row + 2, col + 1)
            data_cell.value = data

    work_book.save(output_file)


if __name__ == '__main__':
    MergeFiles('try_error/',
               output_file='Combined-2.xlsx'
               )
